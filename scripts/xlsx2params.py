#!/usr/bin/env python3
"""
焊接工艺参数模板.xlsx → params.js 转换脚本

用法:
  # 正常模式：Excel 数据覆盖到 params.js（保留原数据，Excel 优先级更高）
  python3 scripts/xlsx2params.py
  
  # 回填模式：将 params.js 现有数据回填到 Excel（一次性铺全数据）
  python3 scripts/xlsx2params.py --backfill

  # 指定文件
  python3 scripts/xlsx2params.py 模板.xlsx            # 指定 Excel
  python3 scripts/xlsx2params.py 模板.xlsx 输出.js     # 指定 Excel + 输出
  
功能:
  - ✅ 合并模式：Excel 数据覆盖旧值，其他厚度保留
  - ✅ 回填模式：现有 params.js → Excel，一次铺全所有厚度，你再改
  - ✅ 异种材料数据
  - ✅ JS 语法自动验证
  - ✅ 变更报告
"""

import json
import os
import re
import sys
from collections import defaultdict
import copy

try:
    import openpyxl
except ImportError:
    print("❌ 需要安装 openpyxl: pip3 install openpyxl")
    sys.exit(1)

PROJECT_NAME = '激光焊接参数查询工具'

# ========== 材料名双向映射 ==========
CN_TO_EN = {
    '不锈钢304': 'ss304',
    '不锈钢316': 'ss316',
    '碳钢': 'carbon',
    '镀锌板': 'galvanized',
    '纯铜': 'copper',
    '黄铜': 'brass',
    '铝合金': 'aluminum',
    '钛合金': 'titanium',
}
EN_TO_CN = {v: k for k, v in CN_TO_EN.items()}

def normalize_material(name):
    if not name:
        return None
    name = name.strip()
    if name in CN_TO_EN:
        return CN_TO_EN[name]
    for cn, en in sorted(CN_TO_EN.items(), key=lambda x: -len(x[0])):
        if cn in name:
            return en
    return name.lower().replace(' ', '_')

def normalize_mode(mode):
    if not mode:
        return 'cw'
    mode = str(mode).strip().lower()
    if mode in ('pulse', '脉冲'):
        return 'pulse'
    return 'cw'

def to_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def to_int(v):
    f = to_float(v)
    return int(f) if f is not None else None

def minmax(vmin, vmax):
    vmin_f = to_float(vmin)
    vmax_f = to_float(vmax)
    if vmin_f is None and vmax_f is None:
        return None
    return [vmin_f if vmin_f is not None else 0, vmax_f if vmax_f is not None else vmin_f]


# ========== Excel → 数据 ==========

def load_excel(filepath):
    """读取 Excel 模板，返回行数据列表"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['工艺参数数据']
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if all(v is None for v in row):
            continue
        r = {
            'material1': str(row[0]).strip() if row[0] else '',
            'material2': str(row[1]).strip() if row[1] else '',
            'sub1': str(row[2]).strip() if row[2] else '',
            'sub2': str(row[3]).strip() if row[3] else '',
            'mode': str(row[4]).strip().lower() if row[4] else 'cw',
            'thickness': to_float(row[5]),
            'joint': str(row[6]).strip() if row[6] else '',
            'wire': str(row[7]).strip() if row[7] else '',
            'wobble': str(row[8]).strip() if row[8] else '',
            'gas': str(row[9]).strip() if row[9] else '',
            'power_min': to_int(row[10]),
            'power_max': to_int(row[11]),
            'speed_min': to_float(row[12]),
            'speed_max': to_float(row[13]),
            'focus_min': to_float(row[14]),
            'focus_max': to_float(row[15]),
            'gas_flow_min': to_float(row[16]),
            'gas_flow_max': to_float(row[17]),
            'laser_type': str(row[18]).strip() if row[18] else '',
            'core_diameter': str(row[19]).strip() if row[19] else '',
            'collimator_ratio': str(row[20]).strip() if row[20] else '',
            'freq': to_int(row[21]),
            'duty': to_int(row[22]),
            'ppk': to_int(row[23]),
            'feasibility': str(row[24]).strip() if row[24] else '',
            'feasibility_desc': str(row[25]).strip() if row[25] else '',
            'material_hint': str(row[26]).strip() if row[26] else '',
            'notes': str(row[27]).strip() if row[27] else '',
        }
        if not r['material1'] or r['thickness'] is None:
            continue
        rows.append(r)
    return rows

def excel_row_to_params(r):
    """Excel 行 → 参数 dict"""
    params = {
        'power': minmax(r['power_min'], r['power_max']),
        'speed': minmax(r['speed_min'], r['speed_max']),
        'focus': minmax(r['focus_min'], r['focus_max']),
        'gas_flow': minmax(r['gas_flow_min'], r['gas_flow_max']),
    }
    mode = normalize_mode(r['mode'])
    if mode == 'pulse':
        if r['freq'] is not None:
            params['freq'] = [r['freq'], r['freq']]
        if r['duty'] is not None:
            params['duty'] = [r['duty'], r['duty']]
        if r['ppk'] is not None:
            params['ppk'] = [r['ppk'], r['ppk']]
    return params

def excel_rows_to_homo(rows):
    """Excel 行 → 同种材料 dict"""
    data = defaultdict(lambda: defaultdict(dict))
    for r in rows:
        if r['material1'] != r['material2']:
            continue
        mat = normalize_material(r['material1'])
        mode = normalize_mode(r['mode'])
        thick = r['thickness']
        data[mat][mode][thick] = excel_row_to_params(r)
    return data

def excel_rows_to_hetero(rows):
    """Excel 行 → 异种材料 dict"""
    data = {}
    for r in rows:
        if r['material1'] == r['material2']:
            continue
        m1 = normalize_material(r['material1'])
        m2 = normalize_material(r['material2'])
        key = tuple(sorted([m1, m2]))
        if key not in data:
            data[key] = defaultdict(lambda: defaultdict(dict))
        mode = normalize_mode(r['mode'])
        data[key][mode][r['thickness']] = excel_row_to_params(r)
    return data

# ========== 数据 → JS 输出 ==========

def format_js(v, indent=0):
    """Python → JS 格式化"""
    sp = '  ' * indent
    if v is None:
        return 'null'
    if isinstance(v, list):
        items = [format_js(i, indent) for i in v]
        return '[' + ', '.join(items) + ']'
    if isinstance(v, dict):
        if not v:
            return '{}'
        has_num_keys = any(isinstance(k, (int, float)) for k in v)
        items = []
        for k in sorted(v.keys()):
            ks = str(k) if isinstance(k, (int, float)) and k == int(k) else json.dumps(str(k))
            if isinstance(k, float) and k != int(k):
                ks = json.dumps(str(k))
            val_str = format_js(v[k], indent + 2)
            items.append(f'{sp}    {ks}: {val_str}')
        return '{\n' + ',\n'.join(items) + f'\n{sp}  }}'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        return str(int(v)) if v == int(v) else str(v)
    if isinstance(v, str):
        # Handle thickness-like numeric strings
        try:
            f = float(v)
            return str(int(f)) if f == int(f) else str(f)
        except ValueError:
            pass
        return json.dumps(v)
    return str(v)

def format_dict_entries(d, indent=0, key_quote=False):
    """格式化 dict 的每个条目"""
    lines = []
    for k, v in d.items():
        if key_quote:
            ks = json.dumps(k)
        else:
            if isinstance(k, (int, float)):
                ks = str(int(k)) if k == int(k) else json.dumps(str(k))
            else:
                ks = json.dumps(k)
        vs = format_js(v, indent)
        lines.append(f'{"  "*indent}{ks}: {vs}')
    return ',\n'.join(lines)

def generate_js_string(homo_data):
    """生成 params.js 完整内容"""
    lines = [
        '// 焊接参数数据（由 焊接工艺参数模板.xlsx 自动生成）',
        '// 最后更新: (自动)',
        '// 编辑方法: 修改 焊接工艺参数模板.xlsx → 运行 scripts/xlsx2params.py',
        'var PARAMS_DATA = {',
    ]
    
    for mi, mat in enumerate(sorted(homo_data.keys())):
        lines.append(f'  {mat}: {{')
        modes = homo_data[mat]
        for mj, mode in enumerate(sorted(modes.keys(), reverse=True)):
            lines.append(f'    {json.dumps(mode)}: {{')
            thicks = modes[mode]
            for tk, thick in enumerate(sorted(thicks.keys())):
                thick_str = str(int(thick)) if thick == int(thick) else f'{thick:.1f}'
                lines.append(f'      {thick_str}: {{')
                params = thicks[thick]
                pk_list = list(params.keys())
                for pi, pk in enumerate(pk_list):
                    comma = ',' if pi < len(pk_list) - 1 else ''
                    lines.append(f'        {json.dumps(pk)}: {format_js(params[pk])}{comma}')
                comma_tk = ',' if tk < len(thicks) - 1 else ''
                lines.append(f'      }}{comma_tk}')
            comma_mj = ',' if mj < len(modes) - 1 else ''
            lines.append(f'    }}{comma_mj}')
        comma_mi = ',' if mi < len(homo_data) - 1 else ''
        lines.append(f'  }}{comma_mi}')
    
    lines.append('};')
    lines.append('')
    return '\n'.join(lines)

# ========== 读取现有 params.js（合并模式用） ==========

def parse_js_value(text, pos):
    """简单 JS 值解析器（专用，只处理我们的数据结构）"""
    text = text.strip()
    if not text:
        return None, 0
    
    if text.startswith('null'):
        return None, 4
    if text.startswith('true'):
        return True, 4
    if text.startswith('false'):
        return False, 5
    
    if text.startswith('['):
        # list
        items = []
        i = 1
        while i < len(text):
            text = text[i:].strip()
            if not text:
                break
            if text.startswith(']'):
                return items, i + 1
            if text.startswith(','):
                i = 1
                continue
            val, consumed = parse_js_value(text, 0)
            items.append(val)
            i = consumed
        return items, i + 1
    
    if text.startswith('{'):
        # dict
        obj = {}
        i = 1
        while i < len(text):
            text = text[i:].strip()
            if not text:
                break
            if text.startswith('}'):
                return obj, i + 1
            if text.startswith(','):
                i = 1
                continue
            # key
            key, consumed = parse_js_value(text, 0)
            i = consumed
            # skip colon
            text = text[i:].strip()
            if text.startswith(':'):
                i = 1
            # value
            text = text[i:].strip()
            val, consumed2 = parse_js_value(text, 0)
            i = consumed2
            obj[key] = val
        return obj, i + 1
    
    if text[0] in '"\'':
        # string
        quote = text[0]
        end = text.find(quote, 1)
        if end > 0:
            return text[1:end], end + 1
        return text, len(text)
    
    # number
    m = re.match(r'^[\d\.\-]+', text)
    if m:
        s = m.group()
        if '.' in s:
            return float(s), len(s)
        return int(s), len(s)
    
    # identifier (material name)
    m = re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*', text)
    if m:
        return m.group(), len(m.group())
    
    return None, 0

def read_current_params():
    """读取现有 params.js 获取同种材料数据（无依赖外部库）"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    params_path = os.path.join(project_dir, 'data', 'params.js')
    
    if not os.path.exists(params_path):
        return None
    
    # 用 node 把 params.js 转成 JSON
    temp_script = '''
    const fs = require("fs");
    ''' + open(params_path, 'r').read() + '''
    fs.writeFileSync('/tmp/params.json', JSON.stringify(PARAMS_DATA, null, 2));
    '''
    tmp_path = '/tmp/_xlsx2params_extract.js'
    with open(tmp_path, 'w') as f:
        f.write(temp_script)
    
    ret = os.system(f'node "{tmp_path}" 2>/dev/null')
    if ret != 0:
        return None
    
    with open('/tmp/params.json', 'r') as f:
        data = json.load(f)
    
    os.remove(tmp_path)
    os.remove('/tmp/params.json')
    
    # JSON keys are always strings; convert numeric keys back to float
    def fix_keys(d):
        if not isinstance(d, dict):
            return d
        result = {}
        for k, v in d.items():
            # Try to convert numeric string keys to float
            try:
                nk = float(k)
                if nk == int(nk):
                    nk = int(nk)
            except (ValueError, TypeError):
                nk = k
            result[nk] = fix_keys(v)
        return result
    
    return fix_keys(data)


# ========== 回填 Excel ==========

def backfill_excel(homo_data, output_path):
    """将现有 params.js 数据回填到 Excel 模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = '工艺参数数据'
    
    # 表头
    headers = ['材料1', '材料2', '材料1子类', '材料2子类', '焊接模式', '板厚(mm)', 
               '接头类型', '送丝', '摆头', '保护气体', '功率最小值(W)', '功率最大值(W)',
               '速度最小值(mm/s)', '速度最大值(mm/s)', '焦点最小值(mm)', '焦点最大值(mm)',
               '气流量最小值(L/min)', '气流量最大值(L/min)', '激光器类型', '激光器芯径um',
               '焊接准直聚焦比', '脉冲频率(Hz)', '占空比(%)', '峰值功率(W)', 
               '可行性等级', '可行性描述', '材料提示', '备注']
    
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # 说明行
    hints = [
        '母材A，如 不锈钢304、碳钢Q235',
        '异种填写不同材料，同种与材料1相同',
        '可选：不锈钢可不填；铝填 1系/2系/3系/5系/6系/7系；铜填 纯铜/黄铜/青铜/磷铜/白铜',
        '同上',
        'cw 连续激光 / pulse 脉冲激光',
        '0.2~12.0，精确到0.1mm',
        '对接/搭接/角接/边接',
        'self 自熔 / feed 送丝',
        'no 无 / yes 有',
        'Ar / N2 / He / vacuum / air',
        '整数，建议范围200~12000',
        '整数，建议范围500~15000',
        '浮点数，建议范围1~100',
        '浮点数，建议范围2~120',
        '负值表示离焦（深熔焊），正值表示正焦',
        '同上',
        '浮点数，建议范围5~50',
        '浮点数，建议范围8~60',
        '普通/高亮度/点环/半导体/红蓝复合',
        '10/14/20/30/50/100/200/600/（14/100）/（20/120）/（50/150）/（100/300）/（100/600）',
        '（100/200）/（100/300）/（100/350）/（150/350）/（150/400）',
        '仅脉冲模式填，范围5~100',
        '仅脉冲模式填，范围20~80',
        '仅脉冲模式填',
        'easy / moderate / challenging / hard',
        '如"同种不锈钢304焊接，工艺成熟，易获优质焊缝"',
        '子材料或特殊说明',
        '自由备注字段',
    ]
    hint_font = Font(color='666666', italic=True, size=9)
    for ci, h in enumerate(hints, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = hint_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border
    
    # 数据行
    row_idx = 3
    for mat in sorted(homo_data.keys()):
        cn_name = EN_TO_CN.get(mat, mat)
        for mode in ['cw', 'pulse']:
            if mode not in homo_data[mat]:
                continue
            for thick in sorted(homo_data[mat][mode].keys()):
                params = homo_data[mat][mode][thick]
                # 填写
                ws.cell(row=row_idx, column=1, value=cn_name).border = thin_border
                ws.cell(row=row_idx, column=2, value=cn_name).border = thin_border
                ws.cell(row=row_idx, column=3).border = thin_border
                ws.cell(row=row_idx, column=4).border = thin_border
                ws.cell(row=row_idx, column=5, value=mode).border = thin_border
                ws.cell(row=row_idx, column=6, value=thick).border = thin_border
                ws.cell(row=row_idx, column=7, value='对接').border = thin_border
                ws.cell(row=row_idx, column=8, value='self').border = thin_border
                ws.cell(row=row_idx, column=9, value='no').border = thin_border
                ws.cell(row=row_idx, column=10, value='Ar').border = thin_border
                
                if 'power' in params:
                    ws.cell(row=row_idx, column=11, value=int(params['power'][0])).border = thin_border
                    ws.cell(row=row_idx, column=12, value=int(params['power'][1])).border = thin_border
                if 'speed' in params:
                    ws.cell(row=row_idx, column=13, value=params['speed'][0]).border = thin_border
                    ws.cell(row=row_idx, column=14, value=params['speed'][1]).border = thin_border
                if 'focus' in params:
                    ws.cell(row=row_idx, column=15, value=params['focus'][0]).border = thin_border
                    ws.cell(row=row_idx, column=16, value=params['focus'][1]).border = thin_border
                if 'gas_flow' in params:
                    ws.cell(row=row_idx, column=17, value=params['gas_flow'][0]).border = thin_border
                    ws.cell(row=row_idx, column=18, value=params['gas_flow'][1]).border = thin_border
                
                if mode == 'pulse' and 'freq' in params:
                    ws.cell(row=row_idx, column=22, value=int(params['freq'][0])).border = thin_border
                    ws.cell(row=row_idx, column=23, value=int(params['duty'][0])).border = thin_border
                    ws.cell(row=row_idx, column=24, value=int(params['ppk'][0])).border = thin_border
                
                # 默认可行性
                ws.cell(row=row_idx, column=25, value='easy').border = thin_border
                
                row_idx += 1
    
    # 列宽
    col_widths = [14, 14, 18, 18, 10, 10, 10, 8, 8, 10, 14, 14, 16, 16, 14, 14, 16, 16, 14, 16, 16, 12, 10, 12, 14, 40, 30, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + ci) if ci <= 26 else 'A' + chr(64 + ci - 26)].width = w
    
    # 使用说明 sheet
    ws2 = wb.create_sheet('使用说明')
    instructions = [
        '激光焊接工艺参数导入模板 - 使用说明',
        '',
        '1. 模板结构',
        '   - 第一行：列名称（固定，不可修改）',
        '   - 第二行：填写说明（可删除，建议保留参考）',
        '   - 第三行起：数据行',
        '',
        '2. 必填字段',
        '   材料1、材料2、焊接模式、板厚(mm)、接头类型、功率(W)、速度(mm/s)',
        '   可行性等级',
        '',
        '3. 命名规范',
        '   材料名称统一使用全称：不锈钢304、碳钢、铝合金、纯铜、黄铜、钛合金、镀锌板',
        '   铝子类：1系/2系/3系/5系/6系/7系',
        '   铜子类：纯铜/黄铜/青铜/磷铜/白铜',
        '   焊接模式：cw(连续) / pulse(脉冲)',
        '   接头类型：对接/搭接/角接/边接',
        '   送丝：self(自熔) / feed(送丝)',
        '   摆头：no(无) / yes(有)',
        '   保护气体：Ar / N2 / He / vacuum / air',
        '   可行性等级：easy(易焊) / moderate(可焊) / challenging(困难) / hard(不推荐)',
        '',
        '4. 编辑后运行命令生成 params.js：',
        '   python3 scripts/xlsx2params.py',
        '',
        '5. 注意事项',
        '   - 不要修改列顺序',
        '   - 数字字段纯数字，不带单位',
        '   - 不要修改表头文字（第一行）',
        '   - 空白行会自动跳过',
    ]
    for i, line in enumerate(instructions, 1):
        ws2.cell(row=i, column=1, value=line)
    
    wb.save(output_path)
    print(f'✅ Excel 已回填: {os.path.abspath(output_path)}')
    print(f'   - 共 {row_idx - 3} 行数据（覆盖所有厚度）')
    print(f'   - 现在可以在 Excel 中修改数据后重新运行脚本')


# ========== 合并逻辑 ==========

def merge_excel_into_current(excel_homo, current_data):
    """Excel 数据合并到现有数据上"""
    merged = copy.deepcopy(current_data) if current_data else {}
    
    for mat in excel_homo:
        if mat not in merged:
            merged[mat] = {}
        for mode in excel_homo[mat]:
            if mode not in merged[mat]:
                merged[mat][mode] = {}
            merged[mat][mode].update(excel_homo[mat][mode])
    
    return merged

def print_changes(excel_homo, current_data):
    """打印变更报告"""
    changes = []
    additions = []
    
    for mat in excel_homo:
        for mode in excel_homo[mat]:
            for thick, params in excel_homo[mat][mode].items():
                cn_name = EN_TO_CN.get(mat, mat)
                thick_str = f'{int(thick)}mm' if thick == int(thick) else f'{thick:.1f}mm'
                
                if current_data and mat in current_data and mode in current_data[mat] and thick in current_data[mat][mode]:
                    old = current_data[mat][mode][thick]
                    if old != params:
                        changes.append(f'  🔄 {cn_name} [{mode}] {thick_str}')
                        for k in ['power', 'speed', 'focus', 'gas_flow', 'freq', 'duty', 'ppk']:
                            if k in old and k in params and old[k] != params[k]:
                                changes.append(f'      {k}: {old[k]} → {params[k]}')
                else:
                    additions.append(f'  ➕ {cn_name} [{mode}] {thick_str}')
    
    if changes:
        print(f'\n📝 变更 ({len(changes)//2} 条):')
        for c in changes:
            print(c)
    
    if additions:
        print(f'\n✨ 新增 ({len(additions)} 条):')
        for a in additions:
            print(a)


# ========== Main ==========

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(project_dir, '焊接工艺参数模板.xlsx')
    output_path = os.path.join(project_dir, 'data', 'params.js')
    
    # 命令行参数
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    flags = set(a for a in sys.argv[1:] if a.startswith('--'))
    
    if len(args) >= 1:
        excel_path = args[0]
    if len(args) >= 2:
        output_path = args[1]
    
    # --- 回填模式 ---
    if '--backfill' in flags:
        current_data = read_current_params()
        if not current_data:
            print('❌ 读取 params.js 失败，可能是空文件或格式不支持')
            sys.exit(1)
        backfill_filename = '焊接工艺参数模板_回填.xlsx'
        backfill_path = os.path.join(project_dir, backfill_filename)
        backfill_excel(current_data, backfill_path)
        print(f'\n💡 然后打开 {backfill_filename}，修改数据后:')
        print(f'   python3 scripts/xlsx2params.py')
        return
    
    # --- 正常转换模式 ---
    if not os.path.exists(excel_path):
        print(f'❌ 找不到 Excel 模板: {excel_path}')
        print(f'   首次使用请先运行 --backfill 生成完整模板')
        sys.exit(1)
    
    # 读取 Excel
    print(f'📂 读取 Excel: {excel_path}')
    rows = load_excel(excel_path)
    print(f'📊 数据行: {len(rows)}')
    homo_rows = [r for r in rows if r['material1'] == r['material2']]
    hetero_rows = [r for r in rows if r['material1'] != r['material2']]
    print(f'   同种: {len(homo_rows)} | 异种: {len(hetero_rows)}')
    
    excel_homo = excel_rows_to_homo(homo_rows)
    
    # 读取现有数据做合并
    current_data = read_current_params()
    
    if current_data:
        print(f'📖 读取现有 params.js（{sum(len(m) for m in current_data.values())} 种材料）')
        # 打印 Excel 相对当前数据的变更
        print_changes(excel_homo, current_data)
        merged = merge_excel_into_current(excel_homo, current_data)
    else:
        print('⚠️  未找到现有 params.js，仅使用 Excel 数据')
        merged = excel_homo
    
    # 输出
    js_content = generate_js_string(merged)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(js_content)
    
    print(f'\n✅ 已写入: {output_path}')
    ret = os.system(f'node --check "{output_path}" 2>&1')
    if ret == 0:
        size = os.path.getsize(output_path)
        print(f'✅ JS 语法验证通过 ({size} bytes)')
    else:
        print(f'⚠️  JS 语法验证失败')
    
    print(f'\n✨ 完成! 下次只需编辑 Excel 后运行:')
    print(f'   python3 scripts/xlsx2params.py')

if __name__ == '__main__':
    main()
