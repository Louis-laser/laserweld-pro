"""
Excel → SQLite 导入工具
把填好的Excel参数表一键导入小程序数据库

用法：
    python3 import_excel.py                   # 导入默认路径
    python3 import_excel.py --file 参数表.xlsx # 指定文件
    python3 import_excel.py --dry-run          # 试跑（只预览，不写入）
"""

import sqlite3
import os
import sys
import argparse

# 默认路径
DEFAULT_EXCEL = os.path.expanduser("~/Desktop/激光焊接工艺参数库_录入模板.xlsx")
DEFAULT_DB = os.path.join(os.path.dirname(__file__), "weld_params.db")

def parse_args():
    parser = argparse.ArgumentParser(description="Excel参数表 → SQLite数据库导入工具")
    parser.add_argument("--file", default=DEFAULT_EXCEL, help="Excel文件路径")
    parser.add_argument("--db", default=DEFAULT_DB, help="目标数据库路径")
    parser.add_argument("--dry-run", action="store_true", help="试跑模式，只预览不写入")
    parser.add_argument("--clear", action="store_true", help="导入前清空旧数据")
    return parser.parse_args()

def read_excel(filepath):
    """读取Excel，返回参数列表"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安装 openpyxl：pip3 install openpyxl")
        sys.exit(1)

    if not os.path.exists(filepath):
        print(f"❌ 未找到文件: {filepath}")
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # 找到表头对应的列索引
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            val = str(val).strip()
            # map field names to column indices
            if "材料1" in val and "材料2" not in val:
                headers["material1"] = col
            elif "材料2" in val:
                headers["material2"] = col
            elif "厚度" in val:
                headers["thickness"] = col
            elif "接头" in val:
                headers["joint_type"] = col
            elif "速度" in val:
                headers["speed"] = col
            elif "功率" in val:
                headers["power"] = col
            elif "离焦" in val:
                headers["defocus"] = col
            elif "保护气" in val:
                headers["shield_gas"] = col
            elif "流量" in val:
                headers["gas_flow"] = col
            elif "工装" in val:
                headers["fixture"] = col
            elif "质量" in val:
                headers["quality"] = col
            elif "备注" in val or "要点" in val:
                headers["note"] = col
            elif "图片" in val:
                headers["image_url"] = col
            elif "视频" in val:
                headers["video_url"] = col

    print(f"  识别的列: {headers}")

    required = ["material1", "material2", "thickness", "joint_type", "speed", "power"]
    missing = [r for r in required if r not in headers]
    if missing:
        print(f"❌ 缺少必要列: {missing}")
        print(f"   请确保Excel表头包含: 材料1, 材料2, 厚度, 接头类型, 焊接速度, 激光功率")
        return None

    params = []
    skip_count = 0
    row_count = 0

    for row in range(header_row + 1, ws.max_row + 1):
        material1 = ws.cell(row=row, column=headers["material1"]).value
        # 跳过空行和示例数据行
        if not material1 or str(material1).strip() in ("", "材料1"):
            continue

        # 跳过示例数据行（数字行号1-5是示例）
        material1 = str(material1).strip()

        try:
            param = {
                "material1": material1,
                "material2": str(ws.cell(row=row, column=headers["material2"]).value or material1).strip(),
                "thickness": float(ws.cell(row=row, column=headers["thickness"]).value or 0),
                "joint_type": str(ws.cell(row=row, column=headers["joint_type"]).value or "").strip(),
                "speed": float(ws.cell(row=row, column=headers["speed"]).value or 0),
                "power": int(float(ws.cell(row=row, column=headers["power"]).value or 0)),
                "defocus": str(ws.cell(row=row, column=headers.get("defocus", 0)).value or "").strip(),
                "shield_gas": str(ws.cell(row=row, column=headers.get("shield_gas", 0)).value or "").strip(),
                "gas_flow": float(ws.cell(row=row, column=headers.get("gas_flow", 0)).value or 0),
                "fixture": str(ws.cell(row=row, column=headers.get("fixture", 0)).value or "").strip(),
                "quality": str(ws.cell(row=row, column=headers.get("quality", 0)).value or "").strip(),
                "note": str(ws.cell(row=row, column=headers.get("note", 0)).value or "").strip(),
                "image_url": str(ws.cell(row=row, column=headers.get("image_url", 0)).value or "").strip(),
                "video_url": str(ws.cell(row=row, column=headers.get("video_url", 0)).value or "").strip(),
            }
            # 跳过空行（关键字段为空）
            if param["thickness"] <= 0 or param["speed"] <= 0 or param["power"] <= 0:
                skip_count += 1
                continue
            params.append(param)
            row_count += 1
        except (ValueError, TypeError):
            skip_count += 1
            continue

    wb.close()
    return params, row_count, skip_count

def import_to_db(params, db_path, clear=False):
    """写入SQLite数据库"""
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    # 建表
    c.execute("""
        CREATE TABLE IF NOT EXISTS weld_params (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            material1 TEXT NOT NULL,
            material2 TEXT NOT NULL,
            thickness REAL NOT NULL,
            joint_type TEXT NOT NULL,
            speed REAL NOT NULL,
            power INTEGER NOT NULL,
            defocus TEXT DEFAULT '',
            shield_gas TEXT DEFAULT '',
            gas_flow REAL DEFAULT 0,
            fixture TEXT DEFAULT '',
            quality TEXT DEFAULT '',
            note TEXT DEFAULT '',
            is_premium INTEGER DEFAULT 0,
            image_url TEXT DEFAULT '',
            video_url TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    if clear:
        c.execute("DELETE FROM weld_params")
        print("  已清空旧数据")

    inserted = 0
    for p in params:
        # 判断是否为高级参数（留给使用者判断，默认所有都可用）
        is_premium = 0
        c.execute("""
            INSERT INTO weld_params 
            (material1, material2, thickness, joint_type, speed, power, 
             defocus, shield_gas, gas_flow, fixture, quality, note,
             is_premium, image_url, video_url)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            p["material1"], p["material2"], p["thickness"], p["joint_type"],
            p["speed"], p["power"], p["defocus"], p["shield_gas"],
            p["gas_flow"], p["fixture"], p["quality"], p["note"],
            is_premium, p["image_url"], p["video_url"]
        ))
        inserted += 1

    conn.commit()
    conn.close()
    return inserted

def main():
    args = parse_args()
    print(f"📂 读取Excel: {args.file}")
    print(f"🗄️  目标数据库: {args.db}")

    result = read_excel(args.file)
    if result is None or result[0] is None:
        print("❌ 导入失败")
        sys.exit(1)

    params, row_count, skip_count = result
    print(f"\n📊 解析结果:")
    print(f"   有效参数: {row_count} 条")
    print(f"   跳过空行: {skip_count} 行")

    if row_count == 0:
        print("⚠️  没有可导入的数据")
        return

    if args.dry_run:
        print("\n🔍 试跑模式 - 以下数据将导入:")
        for p in params[:5]:
            print(f"   {p['material1']} ⟷ {p['material2']}  {p['thickness']}mm  {p['joint_type']}  {p['power']}W")
        if len(params) > 5:
            print(f"   ... 还有 {len(params) - 5} 条")
        print("\n✅ 试跑完成，未写入数据库")
        return

    inserted = import_to_db(params, args.db, clear=args.clear)
    print(f"\n✅ 导入成功！{inserted} 条参数已写入数据库")
    print(f"📍 数据库路径: {args.db}")
    print(f"💡 重启后端服务即可生效: python3 main.py")

if __name__ == "__main__":
    main()
